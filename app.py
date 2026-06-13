import streamlit as st
import pandas as pd
import numpy as np
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
import plotly.express as px

st.set_page_config(
    page_title="TaxTech Auditor RD",
    page_icon="🏛️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── ESTILOS CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stSidebar"] { background: #0f1923; }
    [data-testid="stSidebar"] * { color: #e8edf2 !important; }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stTextInput label,
    [data-testid="stSidebar"] .stSlider label { color: #94a3b8 !important; font-size: 0.78rem !important; }
    [data-testid="stSidebar"] h1 { color: #38bdf8 !important; font-size: 1rem !important; letter-spacing: 0.05em; text-transform: uppercase; }
    .block-container { padding-top: 1.5rem; }
    div[data-testid="metric-container"] { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 1rem; }
    div[data-testid="metric-container"] label { font-size: 0.75rem !important; color: #64748b !important; text-transform: uppercase; letter-spacing: 0.06em; }
    .stTabs [data-baseweb="tab"] { font-size: 0.82rem; padding: 6px 14px; }
    .stTabs [aria-selected="true"] { border-bottom: 2px solid #1e3a5f; font-weight: 700; }
    h1 { color: #1e3a5f !important; }
    h3 { color: #1e3a5f !important; font-size: 1rem !important; }
    
    .tabla-contable { width: 100%; border-collapse: collapse; font-family: 'Calibri', sans-serif; font-size: 0.9rem; margin-bottom: 1rem; }
    .tabla-contable th { border-bottom: 1px solid #000; padding: 8px; text-align: right; font-weight: bold; }
    .tabla-contable th:first-child { text-align: left; }
    .tabla-contable td { padding: 6px 8px; text-align: right; }
    .tabla-contable td:first-child { text-align: left; }
    .tabla-contable .seccion { font-weight: bold; text-align: left; padding-top: 15px; }
    .tabla-contable .total td { border-top: 1px solid #000; border-bottom: 3px double #000; font-weight: bold; }
    .tabla-contable .subtotal td { border-top: 1px solid #ccc; font-weight: bold; }
    .tabla-contable .titulo-anio { text-align: center; font-weight: bold; font-size: 0.95rem; background-color: #f1f5f9; padding: 8px !important; border-bottom: 2px solid #000;}
    
    .tabla-ir2 { width: 100%; border-collapse: collapse; font-family: 'Calibri', sans-serif; font-size: 0.85rem; margin-bottom: 2rem; border: 1px solid #cbd5e1; }
    .tabla-ir2 th { background-color: #1e3a5f; color: white; padding: 10px; text-align: center; font-weight: bold; border: 1px solid #cbd5e1; }
    .tabla-ir2 .header-seccion { background-color: #f1f5f9; font-weight: bold; color: #0f1923; text-align: left; padding: 6px; border: 1px solid #cbd5e1; font-size: 0.9rem; }
    .tabla-ir2 td { padding: 6px; border: 1px solid #cbd5e1; }
    .tabla-ir2 .col-num { width: 6%; text-align: center; font-weight: bold; background-color: #f8fafc; color: #475569; }
    .tabla-ir2 .col-desc { width: 64%; }
    .tabla-ir2 .col-monto { width: 30%; text-align: right; font-family: monospace; font-size:0.95rem;}
    .tabla-ir2 .fila-total td { font-weight: bold; background-color: #e2e8f0; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# PARÁMETROS FISCALES Y CLASIFICACIÓN
# ──────────────────────────────────────────────────────────────────────────────
NATURALEZAS = {'1': 'Debito', '2': 'Credito', '3': 'Credito', '4': 'Credito', '5': 'Debito', '6': 'Debito'}

PALABRAS_CRITICAS_ART287 = {
    'combustible': 'Art. 287 CTRD: Validar NCF válido y medios de pago para crédito ITBIS.',
    'representacion': 'Art. 287 CTRD: Razonabilidad, proporcionalidad y documentación fehaciente.',
    'retribucion': 'Art. 318 / Reg. 139-98: Retribuciones en especie. Validar ISR sustitutivo.',
    'gasto de personal': 'Art. 287 CTRD: Cruce obligatorio con IR-4 (TSS) para admitir deducción.',
    'honorario': 'Art. 309 CTRD: Retención 10% personas físicas / 2% entre jurídicas.',
}

TASA_ITBIS = 0.18
TASA_SFS_PAT = 0.0709; TASA_AFP_PAT = 0.0710
TASA_SRL = 0.0120; TASA_INFOTEP = 0.0100
TASA_SFS_EMP = 0.0304; TASA_AFP_EMP = 0.0287

def fmt_c(val):
    if pd.isna(val) or round(val, 2) == 0: return "-"
    return f"({abs(val):,.0f})" if val < 0 else f"{val:,.0f}"

def es_activo_no_corriente(cod, nombre):
    c = str(cod); n = str(nombre).lower()
    if c.startswith(('15', '16', '17', '18', '19')): return True
    keywords = ['fijo', 'propiedad', 'planta', 'equipo', 'depreciacion', 'edificio', 
                'terreno', 'vehiculo', 'software', 'intangible', 'mejora', 'inversion']
    if any(k in n for k in keywords) and 'gasto' not in n: return True
    return False

def es_pasivo_no_corriente(cod, nombre):
    c = str(cod); n = str(nombre).lower()
    if c.startswith(('22', '23', '24')): return True
    if any(k in n for k in ['largo plazo', 'prestamo bancario', 'bono', 'hipoteca']): return True
    return False

# ──────────────────────────────────────────────────────────────────────────────
# LECTOR DE ARCHIVOS UNIVERSAL (Tolerante a "Falsos Excel" y CSVs)
# ──────────────────────────────────────────────────────────────────────────────
def leer_archivo_robusto(file):
    df = None
    # Intento 1: Como Excel puro
    try:
        file.seek(0)
        df = pd.read_excel(file, header=None)
    except: pass
    
    # Intento 2: Como CSV UTF-8 (El archivo miente con su extensión .xls)
    if df is None or df.empty:
        try:
            file.seek(0)
            df = pd.read_csv(file, header=None, encoding='utf-8', sep=None, engine='python')
        except: pass
        
    # Intento 3: Como CSV Latin1 (Típico en archivos de DGII)
    if df is None or df.empty:
        try:
            file.seek(0)
            df = pd.read_csv(file, header=None, encoding='latin1', sep=None, engine='python')
        except: pass
        
    return df

# ──────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO DE ARCHIVOS (Balanza, 606, 607, TSS)
# ──────────────────────────────────────────────────────────────────────────────
def procesar_balanza(file) -> pd.DataFrame:
    try:
        df_raw = leer_archivo_robusto(file)
        if df_raw is None or df_raw.empty:
            st.warning(f"No se pudo leer el archivo {file.name}. Formato no reconocido.")
            return pd.DataFrame()
        
        header_idx = 0
        for idx, row in df_raw.iterrows():
            row_str = ' '.join([str(x).lower() for x in row.values])
            if ('código' in row_str or 'codigo' in row_str) and ('nombre' in row_str or 'cuenta' in row_str):
                header_idx = idx; break
        
        column_names = df_raw.iloc[header_idx].astype(str).str.lower().str.strip()
        df = df_raw.iloc[header_idx + 1:].reset_index(drop=True)
        
        idx_codigo, idx_cuenta = -1, -1
        indices_debe, indices_haber, indices_balance = [], [], []
        
        for i, col in enumerate(column_names):
            if any(x in col for x in ['código', 'codigo', 'cuenta no']) and idx_codigo == -1: idx_codigo = i
            elif any(x in col for x in ['nombre', 'descripción', 'cuenta']) and 'codigo' not in col and idx_cuenta == -1: idx_cuenta = i
            elif any(x in col for x in ['débito', 'debito', 'debe', 'cargos']): indices_debe.append(i)
            elif any(x in col for x in ['crédito', 'credito', 'haber', 'abonos']): indices_haber.append(i)
            elif any(x in col for x in ['saldo', 'balance', 'final', 'monto']): indices_balance.append(i)
        
        if idx_codigo == -1 or idx_cuenta == -1: 
            st.warning(f"No se detectaron las columnas 'Código' y 'Cuenta' en {file.name}.")
            return pd.DataFrame()
        
        col_dict = {'codigo': df.iloc[:, idx_codigo], 'cuenta': df.iloc[:, idx_cuenta]}
        if indices_debe: col_dict['debito'] = df.iloc[:, indices_debe[-1]]
        if indices_haber: col_dict['credito'] = df.iloc[:, indices_haber[-1]]
        if indices_balance: col_dict['saldo_final'] = df.iloc[:, indices_balance[-1]]
        
        df_clean = pd.DataFrame(col_dict)
        if 'saldo_final' not in df_clean.columns and 'debito' in df_clean.columns:
            df_clean['saldo_final'] = pd.to_numeric(df_clean['debito'], errors='coerce').fillna(0) - pd.to_numeric(df_clean['credito'], errors='coerce').fillna(0)
        
        df_clean['codigo'] = df_clean['codigo'].fillna('').astype(str).str.strip().apply(lambda x: x.split('.')[0] if '.' in x else x)
        df_clean['cuenta'] = df_clean['cuenta'].fillna('').astype(str).str.strip()
        df_clean = df_clean[(df_clean['codigo'] != '') & (~df_clean['codigo'].str.lower().str.contains('total|suma', na=False))]
        
        for col in ['debito', 'credito', 'saldo_final']:
            if col in df_clean.columns:
                df_clean[col] = pd.to_numeric(df_clean[col].astype(str).str.replace(',', '').replace(r'^-*$', '0', regex=True), errors='coerce').fillna(0.0)
                
        return df_clean.reset_index(drop=True)
    except Exception as e: 
        st.error(f"Error crítico al procesar balanza {file.name}: {str(e)}")
        return pd.DataFrame()

def procesar_606_607(file, tipo):
    try:
        df_raw = leer_archivo_robusto(file)
        if df_raw is None or df_raw.empty: return 0.0, 0.0
        
        h_idx = -1
        for idx, row in df_raw.iterrows():
            row_str = ' '.join([str(x).lower() for x in row.values])
            if 'rnc' in row_str and ('ncf' in row_str or 'monto' in row_str):
                h_idx = idx; break
                
        if h_idx == -1: return 0.0, 0.0
        
        df = pd.DataFrame(df_raw.iloc[h_idx + 1:].values, columns=df_raw.iloc[h_idx].astype(str).str.lower().str.strip())
        
        col_monto = next((c for c in df.columns if any(x in str(c) for x in ['total monto facturado', 'monto facturado', 'monto total'])), None)
        
        # Búsqueda selectiva según el tipo de formato
        if tipo == "606":
            col_itbis = next((c for c in df.columns if any(x in str(c) for x in ['itbis por adelantar', 'itbis adelantado'])), None)
            if not col_itbis: col_itbis = next((c for c in df.columns if 'itbis facturado' in str(c)), None)
        else:
            col_itbis = next((c for c in df.columns if any(x in str(c) for x in ['itbis cobrado', 'itbis facturado'])), None)
            
        monto_total = pd.to_numeric(df[col_monto].astype(str).str.replace(',', ''), errors='coerce').fillna(0).sum() if col_monto else 0.0
        itbis_total = pd.to_numeric(df[col_itbis].astype(str).str.replace(',', ''), errors='coerce').fillna(0).sum() if col_itbis else 0.0
        
        return monto_total, itbis_total
    except Exception as e:
        st.warning(f"Error extrayendo {tipo} de {file.name}: {e}")
        return 0.0, 0.0

def procesar_tss(file):
    try:
        df_raw = leer_archivo_robusto(file)
        if df_raw is None or df_raw.empty: return None, 0
            
        h_idx = 0
        for idx, row in df_raw.iterrows():
            if 'cédula' in ' '.join([str(x).lower() for x in row.values]) or 'cedula' in ' '.join([str(x).lower() for x in row.values]):
                h_idx = idx; break
                
        df = pd.DataFrame(df_raw.iloc[h_idx + 1:].values, columns=df_raw.iloc[h_idx].astype(str).str.lower().str.strip())
        
        col_salario = next((c for c in df.columns if 'salario ordinario' in str(c) or 'sueldo' in str(c)), None)
        if not col_salario: return None, 0
        
        df[col_salario] = pd.to_numeric(df[col_salario].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        df_valid = df[df[col_salario] > 0]
        salario_total = df_valid[col_salario].sum()
        
        res = {
            'total_empleados': len(df_valid),
            'nomina_mensual': salario_total,
            'sfs_pat': salario_total * TASA_SFS_PAT,
            'afp_pat': salario_total * TASA_AFP_PAT,
            'srl_pat': salario_total * TASA_SRL,
            'infotep': salario_total * TASA_INFOTEP,
            'sfs_emp': salario_total * TASA_SFS_EMP,
            'afp_emp': salario_total * TASA_AFP_EMP
        }
        res['total_pagar'] = res['sfs_pat'] + res['afp_pat'] + res['srl_pat'] + res['infotep'] + res['sfs_emp'] + res['afp_emp']
        return df_valid, res
    except Exception as e:
        st.error(f"Error procesando plantilla TSS: {e}")
        return None, 0

def generar_plantilla_tss():
    df = pd.DataFrame(columns=[
        "Tipo de Documento", "Cédula / Pasaporte", "Nombres", "Apellidos", "Sexo", 
        "Fecha de Nacimiento", "Salario Ordinario", "Otras Remuneraciones", "Aporte Voluntario"
    ])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return buf.getvalue()

def generar_excel_descargable(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos Exportados')
    return buf.getvalue()

# ──────────────────────────────────────────────────────────────────────────────
# LÓGICA DE AUDITORÍA Y CÁLCULOS
# ──────────────────────────────────────────────────────────────────────────────
def analizar_balanza(df: pd.DataFrame) -> pd.DataFrame:
    alertas_nat, alertas_fisc = [], []
    for row in df.itertuples(index=False):
        cod, nom, saldo = str(row.codigo), str(row.cuenta).lower(), row.saldo_final
        nat_esp = NATURALEZAS.get(cod[0] if cod else '', None)
        if any(x in nom for x in ['acum', 'depreciacion acum', 'deterioro', 'provision', 'amortizacion acum']):
            nat_esp = 'Credito' if nat_esp == 'Debito' else 'Debito'
            
        if nat_esp == 'Debito' and saldo < -1: alertas_nat.append("⚠️ Saldo crédito (nat. débito)")
        elif nat_esp == 'Credito' and saldo > 1: alertas_nat.append("⚠️ Saldo débito (nat. crédito)")
        else: alertas_nat.append("✅ Correcto")
        alertas_fisc.append(next((msg for p, msg in PALABRAS_CRITICAS_ART287.items() if p in nom), ""))
        
    df['validacion_naturaleza'] = alertas_nat
    df['alerta_fiscal'] = alertas_fisc
    return df

def procesar_comparativo(df_act: pd.DataFrame, df_ant: pd.DataFrame) -> pd.DataFrame:
    df_comp = pd.merge(df_act[['codigo', 'cuenta', 'saldo_final']], df_ant[['codigo', 'saldo_final']], on='codigo', how='outer', suffixes=('_Y2', '_Y1')).fillna(0.0)
    df_comp.loc[df_comp['cuenta'] == 0.0, 'cuenta'] = "Cuenta Histórica"
    df_comp['variacion_abs'] = df_comp['saldo_final_Y2'] - df_comp['saldo_final_Y1']
    df_comp['variacion_pct'] = np.where(df_comp['saldo_final_Y1'] != 0, (df_comp['variacion_abs'] / df_comp['saldo_final_Y1'].replace(0, np.nan)), 0)
    return df_comp

def calcular_casillas_ir2(df: pd.DataFrame) -> dict:
    def suma(prefijos):
        mask = df['codigo'].apply(lambda x: any(str(x).startswith(p) for p in prefijos))
        return abs(df.loc[mask, 'saldo_final'].sum())
    total_pasivos = suma(['2'])
    inventario = suma(['13', '130', '131', '132', '133'])
    af1 = suma(['152', '153']); af2 = suma(['154', '155']); af3 = suma(['156', '157', '158'])
    otros_act = suma(['14', '16', '17', '18', '19'])
    saldo_act_fiscal = af1 + af2 + af3 + inventario + otros_act
    patrimonio_fisc = saldo_act_fiscal - total_pasivos
    total_no_monet = af1 + af2 + af3 + inventario
    return {'cas_34': min(max(patrimonio_fisc, 0), total_no_monet)}

# ──────────────────────────────────────────────────────────────────────────────
# EXPORTADOR EXCEL ÍNTEGRO
# ──────────────────────────────────────────────────────────────────────────────
def exportar_reporte_corporativo(empresa, periodo, anio_act, df_comp):
    try:
        wb = openpyxl.Workbook()

        FNT_TITLE  = Font(name="Calibri", size=14, bold=True, color="1F497D")
        FNT_HDR    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        FNT_B      = Font(name="Calibri", size=11, bold=True)
        FNT_N      = Font(name="Calibri", size=11)
        FNT_POS    = Font(name="Calibri", size=11, color="166534")
        FNT_NEG    = Font(name="Calibri", size=11, color="991B1B")
        FILL_HDR   = PatternFill("solid", fgColor="1F497D")
        FILL_SUB   = PatternFill("solid", fgColor="F1F5F9")
        FILL_TOT   = PatternFill("solid", fgColor="E2E8F0")
        FILL_DASH  = PatternFill("solid", fgColor="0F1923")
        THIN       = Side(border_style="thin",   color="CBD5E1")
        B_BTM      = Border(bottom=THIN)
        B_DBL      = Border(top=THIN, bottom=Side(border_style="double", color="000000"))
        FMT_ACC    = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
        FMT_PCT    = '0.00%'
        anio_prev  = int(anio_act) - 1

        def _pct(y2, y1):
            return (y2 - y1) / abs(y1) if y1 != 0 else (1.0 if y2 != 0 else 0.0)

        def format_row(ws, row_n, style_type='normal'):
            for cell in ws[row_n]:
                cell.font = FNT_B if style_type in ['sub', 'sec', 'tot'] else FNT_N
                if style_type == 'sec': cell.fill = FILL_SUB; cell.border = B_BTM
                if style_type == 'tot': cell.fill = FILL_TOT; cell.border = B_DBL
                if isinstance(cell.value, (int, float)):
                    cell.number_format = FMT_PCT if cell.column == 6 else FMT_ACC

        def color_deviation(cell_abs, cell_pct, v2, v1, higher_is_good=True):
            diff = v2 - v1
            good = (diff >= 0) if higher_is_good else (diff <= 0)
            f = FNT_POS if good else FNT_NEG
            cell_abs.font = f; cell_pct.font = f

        def create_sheet_header(ws, title, cols=None):
            default_cols = [
                (45, "Cuenta"), (15, "Nota"),
                (20, f"Año {anio_act}"), (20, f"Año {anio_prev}"),
                (18, "Variación RD$"), (14, "Variación %")
            ]
            cols = cols or default_cols
            for i, (w, _) in enumerate(cols):
                ws.column_dimensions[get_column_letter(i+1)].width = w
            ws["A1"] = empresa.upper()
            ws["A2"] = title
            ws["A3"] = f"Comparativo años fiscales {anio_act} vs {anio_prev} | {periodo}"
            for row_n in range(1, 4): ws[f"A{row_n}"].font = FNT_TITLE
            for i, (_, h) in enumerate(cols, 1):
                c = ws.cell(row=5, column=i, value=h)
                c.font = FNT_HDR; c.fill = FILL_HDR
                c.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.row_dimensions[5].height = 30
            return 6

        # ─────────────────────────────────────────────────────────────────────
        # 1. BALANCE GENERAL
        # ─────────────────────────────────────────────────────────────────────
        ws_bg = wb.active; ws_bg.title = "Balance General"
        r = create_sheet_header(ws_bg, "ESTADO DE SITUACIÓN FINANCIERA")

        def process_section(ws, r_idx, title, prefix, is_current, is_asset, higher_good=True):
            ws.cell(row=r_idx, column=1, value=title)
            format_row(ws, r_idx, 'sec'); r_idx += 1
            tot_y2, tot_y1 = 0, 0
            for _, row in df_comp[df_comp['codigo'].str.startswith(prefix, na=False)].iterrows():
                check_nc = (es_activo_no_corriente if is_asset else es_pasivo_no_corriente)(row['codigo'], row['cuenta'])
                if (is_current and not check_nc) or (not is_current and check_nc):
                    v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                    if is_asset and 'acum' in str(row['cuenta']).lower(): v2, v1 = -v2, -v1
                    if v2 == 0 and v1 == 0: continue
                    tot_y2 += v2; tot_y1 += v1
                    ws.cell(row=r_idx, column=1, value=row['cuenta'].title())
                    ws.cell(row=r_idx, column=3, value=v2); ws.cell(row=r_idx, column=4, value=v1)
                    c_abs = ws.cell(row=r_idx, column=5, value=v2 - v1)
                    c_pct = ws.cell(row=r_idx, column=6, value=_pct(v2, v1)); c_pct.number_format = FMT_PCT
                    color_deviation(c_abs, c_pct, v2, v1, higher_good)
                    format_row(ws, r_idx, 'normal'); r_idx += 1
            lbl = f"Total {title.lower().replace(':','').strip()}"
            ws.cell(row=r_idx, column=1, value=lbl)
            ws.cell(row=r_idx, column=3, value=tot_y2); ws.cell(row=r_idx, column=4, value=tot_y1)
            c_abs = ws.cell(row=r_idx, column=5, value=tot_y2 - tot_y1)
            c_pct = ws.cell(row=r_idx, column=6, value=_pct(tot_y2, tot_y1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, tot_y2, tot_y1, higher_good)
            format_row(ws, r_idx, 'sub'); r_idx += 1
            return r_idx, tot_y2, tot_y1

        r, ac_y2,  ac_y1  = process_section(ws_bg, r, "Activos corrientes:",     '1', True,  True,  True)
        r, anc_y2, anc_y1 = process_section(ws_bg, r, "Activos no corrientes:",  '1', False, True,  True)
        tot_act_y2 = ac_y2 + anc_y2; tot_act_y1 = ac_y1 + anc_y1
        ws_bg.cell(row=r, column=1, value="TOTAL ACTIVOS")
        ws_bg.cell(row=r, column=3, value=tot_act_y2); ws_bg.cell(row=r, column=4, value=tot_act_y1)
        c_abs = ws_bg.cell(row=r, column=5, value=tot_act_y2 - tot_act_y1)
        c_pct = ws_bg.cell(row=r, column=6, value=_pct(tot_act_y2, tot_act_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, tot_act_y2, tot_act_y1)
        format_row(ws_bg, r, 'tot'); r += 2

        r, pc_y2,  pc_y1  = process_section(ws_bg, r, "Pasivos corrientes:",     '2', True,  False, False)
        r, pnc_y2, pnc_y1 = process_section(ws_bg, r, "Pasivos no corrientes:",  '2', False, False, False)

        ws_bg.cell(row=r, column=1, value="Patrimonio:"); format_row(ws_bg, r, 'sec'); r += 1
        pat_y2, pat_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            pat_y2 += v2; pat_y1 += v1
            ws_bg.cell(row=r, column=1, value=row['cuenta'].title())
            ws_bg.cell(row=r, column=3, value=v2); ws_bg.cell(row=r, column=4, value=v1)
            c_abs = ws_bg.cell(row=r, column=5, value=v2 - v1)
            c_pct = ws_bg.cell(row=r, column=6, value=_pct(v2, v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1)
            format_row(ws_bg, r, 'normal'); r += 1

        ws_bg.cell(row=r, column=1, value="Total Patrimonio")
        ws_bg.cell(row=r, column=3, value=pat_y2); ws_bg.cell(row=r, column=4, value=pat_y1)
        c_abs = ws_bg.cell(row=r, column=5, value=pat_y2 - pat_y1)
        c_pct = ws_bg.cell(row=r, column=6, value=_pct(pat_y2, pat_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, pat_y2, pat_y1)
        format_row(ws_bg, r, 'sub'); r += 1

        tot_pap_y2 = pc_y2 + pnc_y2 + pat_y2; tot_pap_y1 = pc_y1 + pnc_y1 + pat_y1
        ws_bg.cell(row=r, column=1, value="TOTAL PASIVOS Y PATRIMONIO")
        ws_bg.cell(row=r, column=3, value=tot_pap_y2); ws_bg.cell(row=r, column=4, value=tot_pap_y1)
        c_abs = ws_bg.cell(row=r, column=5, value=tot_pap_y2 - tot_pap_y1)
        c_pct = ws_bg.cell(row=r, column=6, value=_pct(tot_pap_y2, tot_pap_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, tot_pap_y2, tot_pap_y1)
        format_row(ws_bg, r, 'tot')

        # ─────────────────────────────────────────────────────────────────────
        # 2. ESTADO DE RESULTADOS 
        # ─────────────────────────────────────────────────────────────────────
        ws_er = wb.create_sheet("Estado de Resultados")
        r = create_sheet_header(ws_er, "ESTADO DE RESULTADOS INTEGRALES")

        ws_er.cell(row=r, column=1, value="Ingresos operacionales:"); format_row(ws_er, r, 'sec'); r += 1
        ing_y2, ing_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('4', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            ing_y2 += v2; ing_y1 += v1
            ws_er.cell(row=r, column=1, value=row['cuenta'].title())
            ws_er.cell(row=r, column=3, value=v2); ws_er.cell(row=r, column=4, value=v1)
            c_abs = ws_er.cell(row=r, column=5, value=v2 - v1)
            c_pct = ws_er.cell(row=r, column=6, value=_pct(v2, v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1)
            format_row(ws_er, r, 'normal'); r += 1
        ws_er.cell(row=r, column=1, value="Total Ingresos")
        ws_er.cell(row=r, column=3, value=ing_y2); ws_er.cell(row=r, column=4, value=ing_y1)
        c_abs = ws_er.cell(row=r, column=5, value=ing_y2 - ing_y1)
        c_pct = ws_er.cell(row=r, column=6, value=_pct(ing_y2, ing_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, ing_y2, ing_y1)
        format_row(ws_er, r, 'sub'); r += 1

        ws_er.cell(row=r, column=1, value="Costos de ventas:"); format_row(ws_er, r, 'sec'); r += 1
        cos_y2, cos_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('5', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            cos_y2 += v2; cos_y1 += v1
            ws_er.cell(row=r, column=1, value=row['cuenta'].title())
            ws_er.cell(row=r, column=3, value=-v2); ws_er.cell(row=r, column=4, value=-v1)
            c_abs = ws_er.cell(row=r, column=5, value=-(v2 - v1))
            c_pct = ws_er.cell(row=r, column=6, value=_pct(v2, v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v1, v2) 
            format_row(ws_er, r, 'normal'); r += 1

        ws_er.cell(row=r, column=1, value="Total Costos de Ventas")
        ws_er.cell(row=r, column=3, value=-cos_y2); ws_er.cell(row=r, column=4, value=-cos_y1)
        c_abs = ws_er.cell(row=r, column=5, value=-(cos_y2 - cos_y1))
        c_pct = ws_er.cell(row=r, column=6, value=_pct(cos_y2, cos_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, cos_y1, cos_y2)
        format_row(ws_er, r, 'sub'); r += 1

        ub_y2 = ing_y2 - cos_y2; ub_y1 = ing_y1 - cos_y1
        ws_er.cell(row=r, column=1, value="UTILIDAD BRUTA")
        ws_er.cell(row=r, column=3, value=ub_y2); ws_er.cell(row=r, column=4, value=ub_y1)
        c_abs = ws_er.cell(row=r, column=5, value=ub_y2 - ub_y1)
        c_pct = ws_er.cell(row=r, column=6, value=_pct(ub_y2, ub_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, ub_y2, ub_y1)
        format_row(ws_er, r, 'tot'); r += 1

        ws_er.cell(row=r, column=1, value="Gastos operacionales:"); format_row(ws_er, r, 'sec'); r += 1
        gas_y2, gas_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('6', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            gas_y2 += v2; gas_y1 += v1
            ws_er.cell(row=r, column=1, value=row['cuenta'].title())
            ws_er.cell(row=r, column=3, value=-v2); ws_er.cell(row=r, column=4, value=-v1)
            c_abs = ws_er.cell(row=r, column=5, value=-(v2 - v1))
            c_pct = ws_er.cell(row=r, column=6, value=_pct(v2, v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v1, v2)
            format_row(ws_er, r, 'normal'); r += 1

        ws_er.cell(row=r, column=1, value="Total Gastos Operacionales")
        ws_er.cell(row=r, column=3, value=-gas_y2); ws_er.cell(row=r, column=4, value=-gas_y1)
        c_abs = ws_er.cell(row=r, column=5, value=-(gas_y2 - gas_y1))
        c_pct = ws_er.cell(row=r, column=6, value=_pct(gas_y2, gas_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, gas_y1, gas_y2)
        format_row(ws_er, r, 'sub'); r += 1

        un_y2 = ing_y2 - cos_y2 - gas_y2; un_y1 = ing_y1 - cos_y1 - gas_y1
        ws_er.cell(row=r, column=1, value="UTILIDAD (PÉRDIDA) NETA DEL PERÍODO")
        ws_er.cell(row=r, column=3, value=un_y2); ws_er.cell(row=r, column=4, value=un_y1)
        c_abs = ws_er.cell(row=r, column=5, value=un_y2 - un_y1)
        c_pct = ws_er.cell(row=r, column=6, value=_pct(un_y2, un_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, un_y2, un_y1)
        format_row(ws_er, r, 'tot'); r += 1

        # ─────────────────────────────────────────────────────────────────────
        # ESTADO DE CAMBIOS EN EL PATRIMONIO
        # ─────────────────────────────────────────────────────────────────────
        ws_pat = wb.create_sheet("Patrimonio")
        r = create_sheet_header(ws_pat, "ESTADO DE CAMBIOS EN EL PATRIMONIO")
        
        ws_pat.cell(row=r, column=1, value="Saldos y Movimientos del Período:"); format_row(ws_pat, r, 'sec'); r += 1
        for _, row in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
            v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            ws_pat.cell(row=r, column=1, value=row['cuenta'].title())
            ws_pat.cell(row=r, column=3, value=v2); ws_pat.cell(row=r, column=4, value=v1)
            c_abs = ws_pat.cell(row=r, column=5, value=v2 - v1)
            c_pct = ws_pat.cell(row=r, column=6, value=_pct(v2, v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1)
            format_row(ws_pat, r, 'normal'); r += 1
            
        ws_pat.cell(row=r, column=1, value="TOTAL PATRIMONIO")
        ws_pat.cell(row=r, column=3, value=pat_y2); ws_pat.cell(row=r, column=4, value=pat_y1)
        c_abs = ws_pat.cell(row=r, column=5, value=pat_y2 - pat_y1)
        c_pct = ws_pat.cell(row=r, column=6, value=_pct(pat_y2, pat_y1)); c_pct.number_format = FMT_PCT
        color_deviation(c_abs, c_pct, pat_y2, pat_y1)
        format_row(ws_pat, r, 'tot')

        # ─────────────────────────────────────────────────────────────────────
        # 3. FLUJO DE EFECTIVO
        # ─────────────────────────────────────────────────────────────────────
        ws_fe = wb.create_sheet("Flujo de Efectivo")
        r = create_sheet_header(ws_fe, "ESTADO DE FLUJO DE EFECTIVO (Método Indirecto)")

        def fe_row(ws, r_idx, lbl, v2, v1, higher_good=True, style='normal'):
            ws.cell(row=r_idx, column=1, value=lbl)
            ws.cell(row=r_idx, column=3, value=v2); ws.cell(row=r_idx, column=4, value=v1)
            c_abs = ws.cell(row=r_idx, column=5, value=v2 - v1)
            c_pct = ws.cell(row=r_idx, column=6, value=_pct(v2, v1)); c_pct.number_format = FMT_PCT
            color_deviation(c_abs, c_pct, v2, v1, higher_good)
            format_row(ws, r_idx, style)
            return r_idx + 1

        ws_fe.cell(row=r, column=1, value="I. ACTIVIDADES OPERACIONALES"); format_row(ws_fe, r, 'sec'); r += 1
        r = fe_row(ws_fe, r, "Utilidad neta del período", un_y2, un_y1)
        ws_fe.cell(row=r, column=1, value="Ajustes por partidas no monetarias:"); format_row(ws_fe, r, 'sec'); r += 1
        dep_y2 = abs(df_comp[df_comp['cuenta'].str.lower().str.contains('deprecia', na=False)]['saldo_final_Y2'].sum())
        dep_y1 = abs(df_comp[df_comp['cuenta'].str.lower().str.contains('deprecia', na=False)]['saldo_final_Y1'].sum())
        r = fe_row(ws_fe, r, "(+) Depreciación y amortización", dep_y2, dep_y1)
        ws_fe.cell(row=r, column=1, value="Cambios en capital de trabajo:"); format_row(ws_fe, r, 'sec'); r += 1
        for _, row in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if not es_activo_no_corriente(row['codigo'], row['cuenta']):
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                fe2 = -(v2 - v1); fe1 = 0
                r = fe_row(ws_fe, r, f"  (Aumento)/Disminución: {row['cuenta'].title()}", fe2, fe1, higher_good=False)
        for _, row in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if not es_pasivo_no_corriente(row['codigo'], row['cuenta']):
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                fe2 = v2 - v1; fe1 = 0
                r = fe_row(ws_fe, r, f"  Aumento/(Disminución): {row['cuenta'].title()}", fe2, fe1)
        op_cft_y2 = un_y2 + dep_y2
        op_cft_y1 = un_y1 + dep_y1
        r = fe_row(ws_fe, r, "Efectivo neto de actividades operacionales", op_cft_y2, op_cft_y1, style='tot')

        r += 1
        ws_fe.cell(row=r, column=1, value="II. ACTIVIDADES DE INVERSIÓN"); format_row(ws_fe, r, 'sec'); r += 1
        inv_y2, inv_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if es_activo_no_corriente(row['codigo'], row['cuenta']) and 'acum' not in str(row['cuenta']).lower():
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                fe2 = -(v2 - v1)
                inv_y2 += fe2
                r = fe_row(ws_fe, r, f"  Adquisición/(Venta): {row['cuenta'].title()}", fe2, 0, False)
        r = fe_row(ws_fe, r, "Efectivo neto de actividades de inversión", inv_y2, inv_y1, style='tot')

        r += 1
        ws_fe.cell(row=r, column=1, value="III. ACTIVIDADES DE FINANCIAMIENTO"); format_row(ws_fe, r, 'sec'); r += 1
        fin_y2, fin_y1 = 0, 0
        for _, row in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if es_pasivo_no_corriente(row['codigo'], row['cuenta']):
                v2, v1 = abs(row['saldo_final_Y2']), abs(row['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                fe2 = v2 - v1
                fin_y2 += fe2
                r = fe_row(ws_fe, r, f"  Variación: {row['cuenta'].title()}", fe2, 0)
        r = fe_row(ws_fe, r, "Efectivo neto de actividades de financiamiento", fin_y2, fin_y1, style='tot')

        r += 1
        flujo_neto_y2 = op_cft_y2 + inv_y2 + fin_y2
        r = fe_row(ws_fe, r, "VARIACIÓN NETA EN EFECTIVO Y EQUIVALENTES", flujo_neto_y2, 0, style='tot')

        # ─────────────────────────────────────────────────────────────────────
        # 4. DASHBOARD COMPARATIVO
        # ─────────────────────────────────────────────────────────────────────
        ws_dash = wb.create_sheet("Dashboard", 0)
        ws_dash.sheet_properties.tabColor = "1F497D"
        ws_dash.column_dimensions['A'].width = 28
        for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']: ws_dash.column_dimensions[col].width = 10

        for row_n in range(1, 6):
            for col_n in range(1, 17): ws_dash.cell(row=row_n, column=col_n).fill = FILL_DASH
        ws_dash["B2"] = empresa.upper(); ws_dash["B2"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        ws_dash["B3"] = f"Dashboard Financiero Comparativo — {anio_act} vs {anio_prev}"; ws_dash["B3"].font = Font(name="Calibri", size=11, color="94A3B8")
        ws_dash["B4"] = periodo; ws_dash["B4"].font = Font(name="Calibri", size=10, color="64748B")

        kpi_labels   = [f"Año {anio_act}", f"Año {anio_prev}", "Variación RD$", "Variación %"]
        kpi_concepts = ["Ingresos", "Utilidad Bruta", "Utilidad Neta", "Total Activos", "Total Pasivos", "Patrimonio"]
        kpi_data = [(ing_y2, ing_y1), (ub_y2, ub_y1), (un_y2, un_y1), (tot_act_y2, tot_act_y1), (pc_y2 + pnc_y2, pc_y1 + pnc_y1), (pat_y2, pat_y1)]

        ws_dash.cell(row=7, column=1, value="Indicador").font = FNT_HDR
        ws_dash.cell(row=7, column=1).fill = FILL_HDR
        for i, lbl in enumerate(kpi_labels, 2):
            c = ws_dash.cell(row=7, column=i, value=lbl)
            c.font = FNT_HDR; c.fill = FILL_HDR; c.alignment = Alignment(horizontal="center")

        for idx, (concept, (v2, v1)) in enumerate(zip(kpi_concepts, kpi_data), 8):
            ws_dash.cell(row=idx, column=1, value=concept).font = FNT_B
            ws_dash.cell(row=idx, column=2, value=v2).number_format = FMT_ACC
            ws_dash.cell(row=idx, column=3, value=v1).number_format = FMT_ACC
            c_dif = ws_dash.cell(row=idx, column=4, value=v2 - v1)
            c_dif.number_format = FMT_ACC; c_dif.font = FNT_POS if v2 >= v1 else FNT_NEG
            c_pct_v = ws_dash.cell(row=idx, column=5, value=_pct(v2, v1))
            c_pct_v.number_format = FMT_PCT; c_pct_v.font = FNT_POS if v2 >= v1 else FNT_NEG
            if idx % 2 == 0:
                for col_n in range(1, 6): ws_dash.cell(row=idx, column=col_n).fill = FILL_SUB

        chart1 = BarChart()
        chart1.type = "col"; chart1.grouping = "clustered"
        chart1.title = "P&L Comparativo"; chart1.style = 10
        chart1.y_axis.title = "RD$"; chart1.x_axis.title = "Período"
        chart1.width = 16; chart1.height = 12

        chart1_labels = ["Ingresos", "Costo Ventas", "Utilidad Bruta", "Utilidad Neta"]
        chart1_y2     = [ing_y2, cos_y2, ub_y2, un_y2]
        chart1_y1     = [ing_y1, cos_y1, ub_y1, un_y1]
        for j, (lbl, v2, v1) in enumerate(zip(chart1_labels, chart1_y2, chart1_y1), 8):
            ws_dash.cell(row=15, column=j, value=lbl)
            ws_dash.cell(row=16, column=j, value=v2); ws_dash.cell(row=17, column=j, value=v1)

        data_ref1 = Reference(ws_dash, min_col=8, max_col=11, min_row=16, max_row=17)
        cats_ref1 = Reference(ws_dash, min_col=8, max_col=11, min_row=15)
        chart1.add_data(data_ref1, from_rows=True)
        chart1.set_categories(cats_ref1)
        chart1.series[0].title = SeriesLabel(v=str(anio_act))
        chart1.series[1].title = SeriesLabel(v=str(anio_prev))
        ws_dash.add_chart(chart1, "A20")

        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()
    except Exception as e:
        st.error(f"Error generando Excel Corporativo: {e}")
        return None

# ──────────────────────────────────────────────────────────────────────────────
# GENERADORES DE TABLAS HTML
# ──────────────────────────────────────────────────────────────────────────────
def html_cambios_patrimonio(df_comp, anio):
    html = f"<table class='tabla-contable'><tr><th>Cuentas de Patrimonio</th><th>Saldo Inicial {int(anio)-1}</th><th>Variación</th><th>Saldo Final {anio}</th></tr>"
    pat_y2, pat_y1 = 0, 0
    for _, r in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
        v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
        if v2 == 0 and v1 == 0: continue
        pat_y2 += v2; pat_y1 += v1
        html += f"<tr><td>{r['cuenta'].title()}</td><td>{fmt_c(v1)}</td><td>{fmt_c(v2 - v1)}</td><td>{fmt_c(v2)}</td></tr>"
    html += f"<tr class='total'><td>TOTAL PATRIMONIO</td><td>{fmt_c(pat_y1)}</td><td>{fmt_c(pat_y2 - pat_y1)}</td><td>{fmt_c(pat_y2)}</td></tr></table>"
    return html

def html_estado_resultados(df_comp, anio):
    html = f"<table class='tabla-contable'><tr><th>Conceptos</th><th>Nota</th><th>{anio}</th><th>{int(anio)-1}</th></tr>"
    html += "<tr><td class='seccion' colspan='4'>Ingresos operacionales:</td></tr>"
    ing_y2, ing_y1 = 0, 0
    for _, r in df_comp[df_comp['codigo'].str.startswith('4', na=False)].iterrows():
        v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
        if v2 == 0 and v1 == 0: continue
        ing_y2 += v2; ing_y1 += v1
        html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    html += f"<tr class='subtotal'><td>Total Ingresos</td><td></td><td>{fmt_c(ing_y2)}</td><td>{fmt_c(ing_y1)}</td></tr>"
    
    html += "<tr><td class='seccion' colspan='4'>Costos y gastos operacionales:</td></tr>"
    gas_y2, gas_y1 = 0, 0
    for _, r in df_comp[df_comp['codigo'].str.startswith(('5','6'), na=False)].iterrows():
        v2, v1 = -abs(r['saldo_final_Y2']), -abs(r['saldo_final_Y1'])
        if v2 == 0 and v1 == 0: continue
        gas_y2 += v2; gas_y1 += v1
        html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    html += f"<tr class='subtotal'><td>Total Costos y Gastos</td><td></td><td>{fmt_c(gas_y2)}</td><td>{fmt_c(gas_y1)}</td></tr>"
    html += f"<tr class='total'><td>Utilidad (Pérdida) del Período</td><td></td><td>{fmt_c(ing_y2 + gas_y2)}</td><td>{fmt_c(ing_y1 + gas_y1)}</td></tr></table>"
    return html

def html_balance_general(df_comp, anio, tipo='activo'):
    html = f"<table class='tabla-contable'><tr><th>{tipo.capitalize()}s</th><th>Nota</th><th>{anio}</th><th>{int(anio)-1}</th></tr>"
    tot_c_y2, tot_c_y1, tot_nc_y2, tot_nc_y1 = 0, 0, 0, 0
    
    if tipo == 'activo':
        html += "<tr><td class='seccion' colspan='4'>Activos corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if not es_activo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                tot_c_y2 += v2; tot_c_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total activos corrientes</td><td></td><td>{fmt_c(tot_c_y2)}</td><td>{fmt_c(tot_c_y1)}</td></tr>"
        
        html += "<tr><td class='seccion' colspan='4'>Activos no corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
            if es_activo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if 'acum' in str(r['cuenta']).lower(): v2, v1 = -v2, -v1 
                if v2 == 0 and v1 == 0: continue
                tot_nc_y2 += v2; tot_nc_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total activos no corrientes</td><td></td><td>{fmt_c(tot_nc_y2)}</td><td>{fmt_c(tot_nc_y1)}</td></tr>"

    else:
        html += "<tr><td class='seccion' colspan='4'>Pasivos corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if not es_pasivo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                tot_c_y2 += v2; tot_c_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total pasivos corrientes</td><td></td><td>{fmt_c(tot_c_y2)}</td><td>{fmt_c(tot_c_y1)}</td></tr>"
        
        html += "<tr><td class='seccion' colspan='4'>Pasivos no corrientes:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
            if es_pasivo_no_corriente(r['codigo'], r['cuenta']):
                v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
                if v2 == 0 and v1 == 0: continue
                tot_nc_y2 += v2; tot_nc_y1 += v1
                html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total pasivos no corrientes</td><td></td><td>{fmt_c(tot_nc_y2)}</td><td>{fmt_c(tot_nc_y1)}</td></tr>"
        
        pat_y2, pat_y1 = 0, 0
        html += "<tr><td class='seccion' colspan='4'>Patrimonio:</td></tr>"
        for _, r in df_comp[df_comp['codigo'].str.startswith('3', na=False)].iterrows():
            v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
            if v2 == 0 and v1 == 0: continue
            pat_y2 += v2; pat_y1 += v1
            html += f"<tr><td>{r['cuenta'].title()}</td><td></td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
        html += f"<tr class='subtotal'><td>Total Patrimonio</td><td></td><td>{fmt_c(pat_y2)}</td><td>{fmt_c(pat_y1)}</td></tr>"
    
    gran_tot_y2 = tot_c_y2 + tot_nc_y2 + (pat_y2 if tipo == 'pasivo' else 0)
    gran_tot_y1 = tot_c_y1 + tot_nc_y1 + (pat_y1 if tipo == 'pasivo' else 0)
    titulo_tot = f"Total {tipo.capitalize()}s" if tipo == 'activo' else "Total Pasivos y Patrimonio"
    html += f"<tr class='total'><td>{titulo_tot}</td><td></td><td>{fmt_c(gran_tot_y2)}</td><td>{fmt_c(gran_tot_y1)}</td></tr></table>"
    return html

def html_flujo_hoja_trabajo(df_comp, anio):
    html = f"<table class='tabla-contable'><tr><th>Hoja de Flujo de Efectivo</th><th>{anio}</th><th>{int(anio)-1}</th></tr>"
    html += "<tr><td class='seccion' colspan='3'>Activos</td></tr>"
    for _, r in df_comp[df_comp['codigo'].str.startswith('1', na=False)].iterrows():
        v2, v1 = abs(r['saldo_final_Y2']), abs(r['saldo_final_Y1'])
        if 'acum' in str(r['cuenta']).lower(): v2, v1 = -v2, -v1
        if v2 != 0 or v1 != 0: html += f"<tr><td>{r['cuenta'].title()}</td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    html += "<tr><td class='seccion' colspan='3'>Pasivos</td></tr>"
    for _, r in df_comp[df_comp['codigo'].str.startswith('2', na=False)].iterrows():
        v2, v1 = -abs(r['saldo_final_Y2']), -abs(r['saldo_final_Y1']) 
        if v2 != 0 or v1 != 0: html += f"<tr><td>{r['cuenta'].title()}</td><td>{fmt_c(v2)}</td><td>{fmt_c(v1)}</td></tr>"
    return html + "</table>"

def html_borrador_ir2(df_bal, periodo):
    ingresos = abs(df_bal[df_bal['codigo'].str.startswith('4', na=False)]['saldo_final'].sum())
    costos = abs(df_bal[df_bal['codigo'].str.startswith('5', na=False)]['saldo_final'].sum())
    gastos = abs(df_bal[df_bal['codigo'].str.startswith('6', na=False)]['saldo_final'].sum())
    
    utilidad_neta = ingresos - costos - gastos
    renta_neta_imponible = max(0, utilidad_neta)
    isr_liquidado = renta_neta_imponible * 0.27
    
    activos_totales = abs(df_bal[df_bal['codigo'].str.startswith('1', na=False)]['saldo_final'].sum())
    impuesto_activos = activos_totales * 0.01
    impuesto_mayor = max(isr_liquidado, impuesto_activos)

    html = "<table class='tabla-ir2'>"
    html += f"<tr><th colspan='3'>DECLARACIÓN JURADA ANUAL DEL IMPUESTO SOBRE LA RENTA DE SOCIEDADES (IR-2)<br><span style='font-weight:normal; font-size:0.85rem;'>AÑO FISCAL: {periodo}</span></th></tr>"
    html += "<tr><td colspan='3' class='header-seccion'>I. DETERMINACIÓN DE LA RENTA NETA IMPONIBLE O PÉRDIDA</td></tr>"
    html += f"<tr><td class='col-num'>1</td><td class='col-desc'>Total de Ingresos Brutos</td><td class='col-monto'>RD$ {ingresos:,.2f}</td></tr>"
    html += f"<tr><td class='col-num'>2</td><td class='col-desc'>Menos: Costo de Ventas</td><td class='col-monto' style='color:#dc2626;'>RD$ ({costos:,.2f})</td></tr>"
    html += f"<tr><td class='col-num'>3</td><td class='col-desc'>Menos: Gastos Operacionales y Financieros</td><td class='col-monto' style='color:#dc2626;'>RD$ ({gastos:,.2f})</td></tr>"
    html += f"<tr class='fila-total'><td class='col-num'>4</td><td class='col-desc'>Utilidad Neta antes de Impuestos</td><td class='col-monto'>RD$ {utilidad_neta:,.2f}</td></tr>"
    
    html += "<tr><td colspan='3' class='header-seccion'>II. LIQUIDACIÓN DEL IMPUESTO SOBRE LA RENTA</td></tr>"
    html += f"<tr><td class='col-num'>5</td><td class='col-desc'>Renta Neta Imponible (Base de cálculo)</td><td class='col-monto'>RD$ {renta_neta_imponible:,.2f}</td></tr>"
    html += f"<tr class='fila-total'><td class='col-num'>6</td><td class='col-desc'>Impuesto Liquidado (Tasa del 27%)</td><td class='col-monto'>RD$ {isr_liquidado:,.2f}</td></tr>"
    
    html += "<tr><td colspan='3' class='header-seccion'>III. LIQUIDACIÓN DEL IMPUESTO A LOS ACTIVOS</td></tr>"
    html += f"<tr><td class='col-num'>7</td><td class='col-desc'>Total Activos Imponibles</td><td class='col-monto'>RD$ {activos_totales:,.2f}</td></tr>"
    html += f"<tr class='fila-total'><td class='col-num'>8</td><td class='col-desc'>Impuesto a los Activos (Tasa del 1%)</td><td class='col-monto'>RD$ {impuesto_activos:,.2f}</td></tr>"
    
    html += "<tr><td colspan='3' class='header-seccion'>IV. RESUMEN DE PAGO</td></tr>"
    html += f"<tr class='fila-total' style='background-color:#1e3a5f; color:white;'><td class='col-num' style='background-color:#1e3a5f;'>9</td><td class='col-desc'>IMPUESTO MAYOR A PAGAR</td><td class='col-monto'>RD$ {impuesto_mayor:,.2f}</td></tr>"
    html += "</table>"
    return html

# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR SETUP
# ──────────────────────────────────────────────────────────────────────────────
st.sidebar.title("🏛️ TaxTech Auditor RD")
st.sidebar.markdown("---")
empresa = st.sidebar.text_input("Empresa", value="Empresa de Prueba SRL")
periodo = st.sidebar.text_input("Período", value="Enero - Diciembre 2026")
anio    = st.sidebar.text_input("Año Fiscal", value="2026")

st.sidebar.markdown("---")
st.sidebar.markdown("### Archivos Fiscales Auxiliares")
file_606 = st.sidebar.file_uploader("Subir Formato 606 (Compras)", type=["xlsx", "xls", "csv", "txt"])
file_607 = st.sidebar.file_uploader("Subir Formato 607 (Ventas)", type=["xlsx", "xls", "csv", "txt"])
file_tss = st.sidebar.file_uploader("Subir Plantilla TSS", type=["xlsx", "xls", "csv"])

# ──────────────────────────────────────────────────────────────────────────────
# MAIN SETUP
# ──────────────────────────────────────────────────────────────────────────────
st.title("TaxTech Auditor — Declaraciones Juradas & Estados Financieros")
c_up1, c_up2 = st.columns(2)
with c_up1: uploaded = st.file_uploader("📂 Cargar Balanza (Año Actual)", type=["xlsx", "xls", "csv"])
with c_up2: uploaded_prev = st.file_uploader("📂 Cargar Balanza (Año Anterior)", type=["xlsx", "xls", "csv"])

if uploaded is None:
    st.info("👆 Sube la balanza de comprobación para iniciar.")
    st.stop()

with st.spinner("Procesando datos contables..."):
    df_bal = procesar_balanza(uploaded)
    if df_bal.empty: st.stop()
        
    df_bal = analizar_balanza(df_bal)

    if uploaded_prev:
        df_prev = procesar_balanza(uploaded_prev)
        df_comp = procesar_comparativo(df_bal, df_prev) if not df_prev.empty else pd.DataFrame()
    else:
        df_comp = df_bal.copy()
        df_comp.rename(columns={'saldo_final': 'saldo_final_Y2'}, inplace=True)
        df_comp['saldo_final_Y1'] = 0.0
        df_comp['variacion_abs'] = df_comp['saldo_final_Y2']

# KPIs 
t_ingresos = abs(df_bal[df_bal['codigo'].str.startswith('4', na=False)]['saldo_final'].sum())
t_activos  = abs(df_bal[df_bal['codigo'].str.startswith('1', na=False)]['saldo_final'].sum())
t_costos   = abs(df_bal[df_bal['codigo'].str.startswith('5', na=False)]['saldo_final'].sum())
t_gastos   = abs(df_bal[df_bal['codigo'].str.startswith('6', na=False)]['saldo_final'].sum())
utilidad_neta = t_ingresos - t_costos - t_gastos
ir2_vals = calcular_casillas_ir2(df_bal)

# Módulos Anexos
itbis_606, monto_606 = procesar_606_607(file_606, "606") if file_606 else (0,0)
itbis_607, monto_607 = procesar_606_607(file_607, "607") if file_607 else (0,0)
df_tss, tss_res = procesar_tss(file_tss) if file_tss else (None, None)

st.markdown(f"### 📌 {empresa} — {periodo}")

tab_comp, tab_bg, tab_er, tab_pat, tab_efe, tab_bal, tab_inconsist, tab_art287, tab_ir2, tab_it1, tab_tss, tab_consol = st.tabs([
    "📈 Dashboard", "📊 Balance General", "📉 Estado de Resultados", "💼 Cambios Patrimonio", "🌊 Flujo de Efectivo",
    "📋 Balanza Creada", "🚨 Inconsistencias", "⚖️ Riesgos Art.287", 
    "📝 Borrador IR-2", "🧾 Borrador IT-1", "👥 Auditoría TSS", "🏛️ Consolidado Fiscal"
])

try:
    with tab_comp:
        c1, c2 = st.columns(2)
        with c1: 
            st.metric("Ingresos Año Actual", f"RD$ {t_ingresos:,.0f}")
        with c2:
            excel_bytes = exportar_reporte_corporativo(empresa, periodo, anio, df_comp)
            if excel_bytes: st.download_button("📥 Descargar Reporte Financiero Completo (Excel)", data=excel_bytes, file_name=f"Reporte_{empresa.replace(' ','_')}.xlsx")
        df_chart = pd.DataFrame({'Año': [f"{int(anio)-1}", f"{anio}"], 'Ingresos': [sum(abs(df_comp[df_comp['codigo'].str.startswith('4', na=False)]['saldo_final_Y1'])), t_ingresos], 'Activos': [sum(abs(df_comp[df_comp['codigo'].str.startswith('1', na=False)]['saldo_final_Y1'])), t_activos]})
        st.plotly_chart(px.bar(df_chart, x='Año', y=['Ingresos', 'Activos'], barmode='group'), use_container_width=True)

    with tab_bg:
        c1, c2 = st.columns(2)
        with c1: st.markdown(html_balance_general(df_comp, anio, 'activo'), unsafe_allow_html=True)
        with c2: st.markdown(html_balance_general(df_comp, anio, 'pasivo'), unsafe_allow_html=True)
        
    with tab_er: st.markdown(html_estado_resultados(df_comp, anio), unsafe_allow_html=True)
    with tab_pat: st.markdown(html_cambios_patrimonio(df_comp, anio), unsafe_allow_html=True)
    with tab_efe: st.markdown(html_flujo_hoja_trabajo(df_comp, anio), unsafe_allow_html=True)
    
    with tab_bal: 
        df_show_bal = df_bal[['codigo', 'cuenta', 'saldo_final']]
        st.dataframe(df_show_bal, use_container_width=True)
        st.download_button("📥 Descargar Tabla (Excel)", data=generar_excel_descargable(df_show_bal), file_name="Balanza_Auditoria.xlsx")
        
    with tab_inconsist: 
        df_show_inc = df_bal[~df_bal['validacion_naturaleza'].str.startswith('✅')][['codigo', 'cuenta', 'saldo_final', 'validacion_naturaleza']]
        st.dataframe(df_show_inc, use_container_width=True)
        if not df_show_inc.empty: st.download_button("📥 Descargar Tabla (Excel)", data=generar_excel_descargable(df_show_inc), file_name="Inconsistencias.xlsx", key="btn_inc")
        else: st.success("✅ Sin inconsistencias.")
        
    with tab_art287: 
        df_show_art = df_bal[df_bal['alerta_fiscal'] != ""][['codigo', 'cuenta', 'saldo_final', 'alerta_fiscal']]
        st.dataframe(df_show_art, use_container_width=True)
        if not df_show_art.empty: st.download_button("📥 Descargar Tabla (Excel)", data=generar_excel_descargable(df_show_art), file_name="Riesgos_Art287.xlsx", key="btn_art")
        else: st.success("✅ Sin alertas.")
        
    with tab_ir2: st.markdown(html_borrador_ir2(df_bal, periodo), unsafe_allow_html=True)

    with tab_it1:
        st.markdown("### Pre-Liquidación IT-1 (En base a formatos 606/607 cargados)")
        if not file_606 and not file_607:
            st.warning("Carga los formatos 606 y 607 en la barra lateral para generar el cruce.")
        else:
            col1, col2, col3 = st.columns(3)
            col1.metric("ITBIS Cobrado (607)", f"RD$ {itbis_607:,.2f}")
            col2.metric("ITBIS Adelantado (606)", f"RD$ {itbis_606:,.2f}")
            pago = itbis_607 - itbis_606
            col3.metric("ITBIS a Pagar / (Favor)", f"RD$ {pago:,.2f}")

    with tab_tss:
        st.markdown("### Conciliación de Nómina (TSS)")
        st.download_button("📥 Descargar Plantilla Vacía Autodeterminación", data=generar_plantilla_tss(), file_name="Plantilla_TSS_Vacia.xlsx")
        
        if file_tss and tss_res:
            c1, c2 = st.columns(2)
            c1.metric("Empleados Validados", tss_res['total_empleados'])
            c1.metric("Monto Total Nómina", f"RD$ {tss_res['nomina_mensual']:,.2f}")
            c2.metric("Aportes Patronales Estimados", f"RD$ {(tss_res['sfs_pat'] + tss_res['afp_pat'] + tss_res['srl_pat'] + tss_res['infotep']):,.2f}")
            c2.metric("Retenciones Empleados", f"RD$ {(tss_res['sfs_emp'] + tss_res['afp_emp']):,.2f}")
            st.dataframe(df_tss, use_container_width=True)
            st.download_button("📥 Descargar Tabla TSS (Excel)", data=generar_excel_descargable(df_tss), file_name="Auditoria_TSS.xlsx")
        else:
            st.info("Sube la plantilla de Autodeterminación TSS para auditar el cálculo de retenciones y aportes.")

    with tab_consol:
        st.markdown("### Resumen Fiscal Consolidado")
        isr_est = max(0, utilidad_neta) * 0.27
        df_consol = pd.DataFrame([
            ("IT-1", "ITBIS Mensual (Cruce 606/607)", f"RD$ {(itbis_607 - itbis_606):,.2f}", "Calculado"),
            ("IR-3", "Seguridad Social (TSS Estimado)", f"RD$ {tss_res['total_pagar'] if tss_res else 0:,.2f}", "Calculado" if tss_res else "Falta Plantilla"),
            ("IR-2", "Impuesto Renta Estimado", f"RD$ {isr_est:,.2f}", "Base Balanza"),
        ], columns=["Formulario", "Concepto", "Monto Estimado", "Estado"])
        st.dataframe(df_consol, use_container_width=True, hide_index=True)
        st.download_button("📥 Descargar Consolidado (Excel)", data=generar_excel_descargable(df_consol), file_name="Consolidado_Fiscal.xlsx")

except Exception as e:
    st.error(f"Error al renderizar: {e}")